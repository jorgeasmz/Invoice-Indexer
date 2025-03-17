import openpyxl
import pandas as pd
import os
from datetime import datetime
import logging
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelWriter:
    """Genera archivos Excel a partir de datos extraídos de facturas"""
    
    def write_invoice_to_excel(self, invoice_data, output_path=None, ocr_results=None, image=None):
        """
        Genera un archivo Excel con la información de una única factura
        
        Args:
            invoice_data: Diccionario con datos extraídos de la factura
            output_path: Ruta donde guardar el archivo Excel (opcional)
            ocr_results: Resultados del OCR (opcional)
            image: Imagen original de la factura (opcional)
            
        Returns:
            str: Ruta al archivo Excel generado
        """
        # Generar nombre de archivo si no se proporciona
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            invoice_num = invoice_data.get('invoice_number', 'unknown')
            output_path = f"factura_{invoice_num}_{timestamp}.xlsx"
        
        logger.info(f"Generando archivo Excel para factura individual: {output_path}")
        
        # Crear un workbook de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos Extraídos"
        
        # Crear encabezados como campos genéricos para facturas
        headers = [
            'Núm. Factura', 
            'Fecha', 
            'Proveedor',
            'NIF/CIF Proveedor', 
            'Cliente', 
            'NIF/CIF Cliente',
            'Concepto',
            'Base Imponible',
            'Tipo IVA',
            'Importe IVA',
            'Total Factura',
            'Forma de Pago',
            'Fecha Vencimiento'
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Escribir datos extraídos
        row_data = [
            invoice_data.get('invoice_number', ''),
            invoice_data.get('date', ''),
            invoice_data.get('supplier_name', ''),
            invoice_data.get('supplier_id', ''),
            invoice_data.get('client_name', ''),
            invoice_data.get('client_id', ''),
            self._get_concept_text(invoice_data),
            invoice_data.get('base_amount', ''),
            invoice_data.get('vat_rate', ''),
            invoice_data.get('vat_amount', ''),
            invoice_data.get('total', ''),
            invoice_data.get('payment_method', ''),
            invoice_data.get('due_date', '')
        ]
        
        # Escribir datos en la fila 2
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Ajustar ancho de las columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # Añadir nota de procesamiento
        footer_row = 4
        ws.cell(row=footer_row, column=1).value = f"Procesado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8)
        ws.merge_cells(f'A{footer_row}:C{footer_row}')
        
        # Guardar
        wb.save(output_path)
        logger.info(f"Archivo Excel generado exitosamente: {output_path}")
        return output_path
    
    def write_multiple_invoices(self, invoice_list, output_path=None):
        """
        Genera un único archivo Excel con datos de múltiples facturas
        
        Args:
            invoice_list: Lista de diccionarios, cada uno con datos de una factura
            output_path: Ruta donde guardar el archivo Excel
            
        Returns:
            str: Ruta al archivo Excel generado
        """
        # Generar nombre de archivo si no se proporciona
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"facturas_procesadas_{timestamp}.xlsx"
        
        logger.info(f"Generando archivo Excel para {len(invoice_list)} facturas: {output_path}")
        
        # Crear un workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas Procesadas"
        
        # Crear encabezados como campos genéricos para facturas
        headers = [
            'Núm. Factura', 
            'Fecha', 
            'Proveedor',
            'NIF/CIF Proveedor', 
            'Cliente', 
            'NIF/CIF Cliente',
            'Concepto',
            'Base Imponible',
            'Tipo IVA',
            'Importe IVA',
            'Total Factura',
            'Forma de Pago',
            'Fecha Vencimiento',
            'Archivo Original'
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Escribir cada factura en una fila
        for row_num, invoice_data in enumerate(invoice_list, 2):
            row_data = [
                invoice_data.get('invoice_number', ''),
                invoice_data.get('date', ''),
                invoice_data.get('supplier_name', ''),
                invoice_data.get('supplier_id', ''),
                invoice_data.get('client_name', ''),
                invoice_data.get('client_id', ''),
                self._get_concept_text(invoice_data),
                invoice_data.get('base_amount', ''),
                invoice_data.get('vat_rate', ''),
                invoice_data.get('vat_amount', ''),
                invoice_data.get('total', ''),
                invoice_data.get('payment_method', ''),
                invoice_data.get('due_date', ''),
                invoice_data.get('metadata', {}).get('file', '') if 'metadata' in invoice_data else ''
            ]
            
            # Escribir datos en la fila
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Aplicar filtro a la tabla
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(invoice_list) + 1}"
        
        # Ajustar ancho de las columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # Añadir información de procesamiento
        footer_row = len(invoice_list) + 3
        ws.cell(row=footer_row, column=1).value = f"Procesado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Total facturas: {len(invoice_list)}"
        ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8)
        ws.merge_cells(f'A{footer_row}:D{footer_row}')
        
        # Guardar
        wb.save(output_path)
        logger.info(f"Archivo Excel con múltiples facturas generado exitosamente: {output_path}")
        return output_path
    
    def _get_concept_text(self, invoice_data):
        """
        Extrae el concepto principal de la factura basado en los ítems
        """
        if 'items' in invoice_data and invoice_data['items']:
            # Usar el primer ítem o concatenar si hay varios
            if len(invoice_data['items']) == 1:
                return invoice_data['items'][0].get('description', '')
            else:
                # Si hay múltiples ítems, mostrar el primero con indicador
                return f"{invoice_data['items'][0].get('description', '')} y {len(invoice_data['items'])-1} conceptos más"
        return ''